from openpyxl import load_workbook
import os
import string

#this object will have the data for one file
class Data:

    def __init__(self, name):
        self.name = name
        self.dataColumn =[]
        self.wb = load_workbook(filename=self.name)

    def row_range(self):
        row_count = 0
        #wb = load_workbook(filename=self.name)
        sheet_ranges = self.wb['Sheet1']
        for x in range (1,10):
            if(sheet_ranges['A' + str(x)].value is None):
                x = 10
            else:
                row_count = row_count + 1
        return row_count

    def column_range(self):
        column_count = 0
        #wb = load_workbook(filename=self.name)
        sheet_ranges = self.wb['Sheet1']
        letters = string.ascii_uppercase
        for x in letters:
            if (sheet_ranges[x + str(1)].value is None):
                return column_count
            else:
                column_count = column_count + 1
        return column_count


    def addColumn(self):
        templist = []
        #wb = load_workbook(filename=self.name)
        sheet_ranges = self.wb['Sheet1']
        column = self.column_range()
        row = self.row_range()
        for x in range (1,column+1):
            for y in range (1, row+1):
                letters = string.ascii_uppercase
                templist.append(sheet_ranges[letters[x-1] + str(y)].value)
            self.dataColumn.append(templist)
            templist = []


def getFileName():
    names = []
    for file in os.listdir("C:/Users/ss5399/Desktop/datatest"):
        if file.endswith(".xlsx"):
            names.append(file)
    return names


def main():
    name = getFileName()
    one = Data(name[0])
    one.addColumn()
    print(one.dataColumn[1])
    two = Data(name[1])
    two.addColumn()
    print(two.dataColumn[1])


if __name__ == "__main__":
    main()
